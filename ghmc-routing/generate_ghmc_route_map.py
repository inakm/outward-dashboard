#!/usr/bin/env python3
"""
Generate the SkyLimit delivery-route -> GHMC administrative map.

Outputs (into --output-dir):
  1. ghmc_route_map.xlsx  - formatted Excel workbook (dark navy header, zebra
                            striping, visible gridlines, auto column width,
                            auto-filter, frozen header, print-ready).
  2. ghmc_route_map.json  - machine-readable mirror of the same rows.
Plus a third file at the repo root:
  3. ghmc-route-map.js    - the same data bundled as a plain script
                            (window.GHMC_ROUTE_MAP = ...) so the dashboard's
                            route map works when opened from the file system.

Routing logic uses the Adjacent-Route Overlap Strategy from the Old Bowenpally
hub (Secunderabad): four diagonal quadrants split on the hub's lat/lng lines,
plus a separate outside-GHMC route. Vehicles may swing only into adjacent
quadrants (never the opposite one) so a single vehicle can cover 1-2 routes
without wasteful cross-city runs.
    R1  North-East (NE)  -> Secunderabad hub belt (Bowenpally) + Alwal/Kompally/
                            Dundigal + Kapra/Keesara north-east flank
    R2  North-West (NW)  -> Quthbullapur north-west wings (Jeedimetla,
                            Gajularamaram, Nizampet, Chintal) + Kukatpally
                            north-west (Kukatpally, Miyapur, Allwyn) +
                            NW fringe (Patancheru, Ameenpur)
    R3  South-East (SE)  -> Kapra/Uppal south-east (Uppal, Boduppal, Nacharam,
                            Malkajgiri, Moula Ali, Tarnaka, Mettuguda) +
                            Secunderabad south wing (Kavadiguda, Musheerabad,
                            Amberpet) + Charminar south-east (Chandrayangutta,
                            Yakutpura, Santoshnagar, Malakpet, Moosarambagh) +
                            LB Nagar zone (Nagole, Saroornagar, LB Nagar,
                            Hayathnagar)
    R4  South-West (SW)  -> Serilingampally south-west (Serilingampally,
                            Madhapur, Narsingi) + Moosapet + Ameerpet +
                            west Khairatabad (Jubilee Hills, Yousufguda,
                            Borabanda) + Old City west (Falaknuma, Bahadurpura,
                            Jangammet, Charminar) + Golconda/Rajendranagar +
                            central Khairatabad (Khairatabad, Mehdipatnam,
                            Masab Tank)
    R5  Outside GHMC     -> outer non-GHMC boundary regions (Sangareddy,
                            Medchal rural, Shamshabad, Ghatkesar,
                            Yadadri Bhuvanagiri)

Adjacency (swing allowed) is diagonal: NE<->NW, NE<->SE, SW<->NW, SW<->SE.
Opposite (blocked) pairs: NE<->SW and NW<->SE.

GHMC circle/ward references follow the 2026 GHMC re-organisation
(12 zones / 60 circles / 300 wards, G.O.Ms.No.292 dt. 24-12-2025).
Zone labels follow the client routing taxonomy, not the 2026 zone names.
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# 1. Canonical routing data                                                  #
# --------------------------------------------------------------------------- #
# Each circle = (circle_name, logistics_note, [(ward_no, ward_name), ...])
# ward_no = None means "not a GHMC ward" (R5 only).

DATA = {
    "R1": {
        "name": "North-East (NE) Route",
        "zones": [
            {
                "zone": "Secunderabad Zone - Hub Belt",
                "circles": [
                    (
                        "Bowenpally",
                        "Hub-adjacent (Bowenpally): sort and load first batch; dispatch before 08:30 AM; 2-4 km short-haul loops.",
                        [
                            (196, "Monda Market"),
                            (260, "Fateh Nagar"),
                            (261, "Prakash Nagar"),
                            (262, "Old Bowenpally"),
                            (264, "Hasmathpet"),
                        ],
                    ),
                ],
            },
            {
                "zone": "Kukatpally/Quthbullapur Zone - North-East Wing",
                "circles": [
                    (
                        "Alwal",
                        "North-east wing on NH44 flank; batch with Quthbullapur belt; morning dispatch.",
                        [
                            (190, "Turkapally"),
                            (191, "Macha Bollaram"),
                            (192, "Temple Alwal"),
                            (193, "Venkatapuram"),
                            (194, "Bhudevi Nagar"),
                            (195, "Kanajiguda"),
                        ],
                    ),
                    (
                        "Kompally",
                        "Quthbullapur north-east (Kompally, Doolapally); gated layouts; morning dispatch with Quthbullapur belt.",
                        [
                            (288, "Kompally"),
                            (289, "Doolapally"),
                            (290, "Subhash Nagar"),
                            (292, "Saibaba Nagar"),
                        ],
                    ),
                    (
                        "Dundigal",
                        "Quthbullapur far north-east (Dundigal, Bowrampet); drone/air-force flank; truck or LCV; morning only.",
                        [
                            (294, "Bahadurpally"),
                            (295, "Bowrampet"),
                            (296, "Dundigal"),
                        ],
                    ),
                ],
            },
            {
                "zone": "Kapra Zone - North-East Belt",
                "circles": [
                    (
                        "Kapra",
                        "ECIL & A.S. Rao Nagar belt; security gate protocols at ECIL; 10 AM-3 PM window.",
                        [
                            (13, "Vampuguda"),
                            (14, "Kapra"),
                            (15, "Dr AS Rao Nagar"),
                            (16, "Kushaiguda"),
                            (17, "Cherlapally"),
                        ],
                    ),
                    (
                        "Keesara",
                        "North-east fringe (Keesara, Dammaiguda, Yapral); one loop via Nagaram road; morning dispatch.",
                        [
                            (1, "Keesara"),
                            (2, "Chandrapuri Colony"),
                            (3, "Jawahar Nagar"),
                            (4, "Dammaiguda"),
                            (189, "Yapral"),
                        ],
                    ),
                ],
            },
        ],
    },
    "R2": {
        "name": "North-West (NW) Route",
        "zones": [
            {
                "zone": "Kukatpally/Quthbullapur Zone - North-West Wings",
                "circles": [
                    (
                        "Jeedimetla",
                        "IDA Jeedimetla industrial belt; truck dock 10 AM-4 PM; gate passes required.",
                        [
                            (284, "Ganesh Nagar"),
                            (285, "Padma Nagar"),
                            (286, "Quthbullapur"),
                            (287, "Pet Basheerabad"),
                        ],
                    ),
                    (
                        "Gajularamaram",
                        "Pharma/industrial corridor; gate-based drop-offs; combine Alwal loop.",
                        [
                            (277, "Mahadevpuram"),
                            (278, "Gajularamaram"),
                            (291, "Shapur Nagar"),
                            (293, "Suraram"),
                        ],
                    ),
                    (
                        "Nizampet",
                        "Residential north-west wing (Bachupally); off-peak after 10 AM; combine Jeedimetla return.",
                        [
                            (273, "Nizampet"),
                            (274, "Bachupally"),
                            (275, "Bhandari Layout"),
                            (276, "Pragathi Nagar"),
                        ],
                    ),
                    (
                        "Chintal",
                        "Quthbullapur industrial north-west (Chintal, Jagathgiri Gutta); truck-friendly; combine Jeedimetla.",
                        [
                            (279, "Rodamestri Nagar"),
                            (280, "Jagathgiri Gutta"),
                            (281, "Ranga Reddy Nagar"),
                            (282, "Chintal"),
                            (283, "Giri Nagar"),
                        ],
                    ),
                ],
            },
            {
                "zone": "Kukatpally Zone - North-West Wing",
                "circles": [
                    (
                        "Kukatpally",
                        "KPHB high-density residential; small vehicles; 6-10 AM window.",
                        [
                            (249, "Kukatpally"),
                            (250, "Balaji Nagar"),
                            (251, "Vasanth Nagar"),
                            (252, "KPHB Colony"),
                            (253, "Kaithalapur"),
                            (254, "Gayatri Nagar"),
                        ],
                    ),
                    (
                        "Miyapur",
                        "Chandanagar/Hafeezpet/Miyapur on NH65 corridor; combine Kukatpally south; mid-day.",
                        [
                            (236, "Hafeezpet"),
                            (237, "Madeenaguda"),
                            (238, "Chanda Nagar"),
                            (239, "Deepthisri Nagar"),
                            (240, "Miyapur"),
                            (241, "Maktha Mahabubpet"),
                        ],
                    ),
                    (
                        "Allwyn Colony",
                        "Kukatpally south residential enclave; narrow lanes; 2W last mile; combine Kukatpally.",
                        [
                            (243, "Hyder Nagar"),
                            (244, "Bhagya Nagar Colony"),
                            (245, "Shamshiguda"),
                            (246, "Allwyn Colony"),
                            (247, "Vivekananda Nagar Colony"),
                            (248, "Venkateshwara Nagar"),
                        ],
                    ),
                ],
            },
            {
                "zone": "Serilingampally Zone - North-West Fringe",
                "circles": [
                    (
                        "Patancheru",
                        "West fringe on NH65 (Tellapur, Patancheruvu); industrial belt; truck-friendly; plan toll.",
                        [
                            (263, "Tellapur"),
                            (265, "Muthangi"),
                            (266, "Patancheruvu"),
                            (267, "JP Colony"),
                        ],
                    ),
                    (
                        "Ameenpur",
                        "West fringe (RC Puram, Beeramguda, Ameenpur); NH65 corridor; combine Patancheru; mid-day.",
                        [
                            (268, "Ramachandrapuram (RC Puram)"),
                            (269, "Bharathi Nagar"),
                            (270, "Beeramguda"),
                            (271, "Ameenpur"),
                            (272, "Bollaram"),
                        ],
                    ),
                ],
            },
        ],
    },
    "R3": {
        "name": "South-East (SE) Route",
        "zones": [
            {
                "zone": "Kapra/Uppal Zone - South-East Belt",
                "circles": [
                    (
                        "Uppal",
                        "IT/industrial mix; IDA Uppal truck gates open till 4 PM; combine Boduppal.",
                        [
                            (23, "Chilkanagar"),
                            (24, "Beerappagadda"),
                            (25, "Habsiguda"),
                            (26, "Ramanthapur"),
                            (27, "Venkat Reddy Nagar"),
                            (28, "Uppal"),
                        ],
                    ),
                    (
                        "Boduppal",
                        "Fast-growing corridor; wide roads; high drop density; multiple quick stops.",
                        [
                            (9, "Medipally"),
                            (10, "Peerzadiguda"),
                            (11, "Boduppal"),
                            (12, "Chengicherla"),
                        ],
                    ),
                    (
                        "Nacharam",
                        "IDA Nacharam industrial; heavy-vehicle restrictions on ORR feeder 6-9 PM; run mid-day.",
                        [
                            (18, "Shakthi Sai Nagar"),
                            (19, "H.B. Colony"),
                            (20, "Mallapur"),
                            (21, "Nacharam"),
                            (22, "HMT Nagar"),
                        ],
                    ),
                    (
                        "Malkajgiri",
                        "Railway colony + gated societies; combine Kapra/Nacharam loop; early start.",
                        [
                            (180, "East Anandbagh"),
                            (181, "Mirjalguda"),
                            (182, "Goutham Nagar"),
                            (183, "Malkajgiri"),
                        ],
                    ),
                    (
                        "Moula Ali",
                        "East-central (Neredmet, Moula Ali, Vinayak Nagar); combine Malkajgiri & Sainikpuri loop.",
                        [
                            (184, "Balram Nagar"),
                            (185, "Vinayak Nagar"),
                            (186, "Moula Ali"),
                            (187, "Kakatiya Nagar"),
                            (188, "Neredmet"),
                        ],
                    ),
                    (
                        "Tarnaka",
                        "Defence-research corridor (DRDO/ECIL flank); gate-entry paperwork; allow 15-min dwell per stop.",
                        [
                            (172, "Boudha Nagar"),
                            (173, "Tarnaka"),
                            (174, "Seethaphalmandi"),
                            (175, "Chilkalguda"),
                        ],
                    ),
                    (
                        "Mettuguda",
                        "Railway-colony lanes; narrow roads; 2W/LCV only; morning slots.",
                        [
                            (176, "Mettuguda"),
                            (177, "Lalapet"),
                            (178, "North Lalaguda"),
                            (179, "Addagutta"),
                        ],
                    ),
                ],
            },
            {
                "zone": "Secunderabad Zone - South-East Wing",
                "circles": [
                    (
                        "Kavadiguda",
                        "Secunderabad core incl. Cantonment links; avoid Parade Grounds & Clock Tower rush 5-7 PM; LCV/2W only.",
                        [
                            (165, "Gandhi Nagar"),
                            (166, "Kavadiguda"),
                            (167, "Bakaram"),
                            (168, "Bholakpur"),
                            (197, "Padmarao Nagar"),
                            (198, "Bansilalpet"),
                            (199, "Ramgopalpet"),
                        ],
                    ),
                    (
                        "Musheerabad",
                        "Dense residential + market belt; early 6-10 AM window; pair with Amberpet run.",
                        [
                            (163, "Adikmet"),
                            (164, "Bagh Lingampally"),
                            (169, "Musheerabad"),
                            (170, "Ramnagar"),
                            (171, "Bapuji Nagar"),
                        ],
                    ),
                    (
                        "Amberpet",
                        "OU/university belt + market; schedule after 11 AM; combine Kavadiguda/Musheerabad belt.",
                        [
                            (155, "Barkatpura"),
                            (156, "Kachiguda"),
                            (157, "Golnaka"),
                            (158, "Patel Nagar"),
                            (159, "Amberpet"),
                            (160, "Bagh Amberpet"),
                            (161, "Tilak Nagar"),
                            (162, "Nallakunta"),
                        ],
                    ),
                ],
            },
            {
                "zone": "Charminar Zone - South-East Belt",
                "circles": [
                    (
                        "Chandrayangutta",
                        "Barkas belt; steep lanes; small vehicles; morning window.",
                        [
                            (68, "Bandlaguda"),
                            (69, "Noori Nagar"),
                            (70, "Barkas"),
                            (71, "Kanchanbagh"),
                            (72, "Chandrayangutta"),
                        ],
                    ),
                    (
                        "Yakutpura",
                        "Charminar market spill; pedestrian-heavy; early morning drops; hand-cart final mile.",
                        [
                            (78, "Gowlipura"),
                            (79, "Talab Chanchalam"),
                            (80, "Yakutpura"),
                            (81, "Dabeerpura"),
                            (82, "Rein Bazar"),
                            (83, "Madannapet"),
                        ],
                    ),
                    (
                        "Santoshnagar",
                        "Mixed density mid-commercial; schedule 10 AM-1 PM.",
                        [
                            (84, "Bhanu Nagar"),
                            (85, "Santosh Nagar"),
                            (86, "IS Sadan"),
                            (87, "Saraswati Nagar"),
                        ],
                    ),
                    (
                        "Malakpet",
                        "Commercial mid-belt; combine Moosarambagh; post-11 AM.",
                        [
                            (88, "Saidabad"),
                            (89, "Asmangadh"),
                            (93, "Akberbagh"),
                            (94, "Chawani"),
                        ],
                    ),
                    (
                        "Moosarambagh",
                        "Ring-road access; quick drops; combine Malakpet.",
                        [
                            (90, "Moosarambagh"),
                            (91, "Old Malakpet"),
                            (92, "MCH Colony"),
                            (95, "Kala Dera"),
                            (96, "Azampura"),
                        ],
                    ),
                ],
            },
            {
                "zone": "L.B. Nagar Zone (South-East)",
                "circles": [
                    (
                        "Nagole",
                        "South-east gateway; start after 9 AM; combine Saroornagar and LB Nagar.",
                        [
                            (29, "Nagole"),
                            (45, "Mansoorabad"),
                            (46, "GSI"),
                            (47, "Lecturers Colony"),
                            (51, "Kuntloor"),
                            (52, "Pedda Amberpet"),
                        ],
                    ),
                    (
                        "Saroornagar",
                        "High-density residential/market; avoid 9-11 AM; schedule 11 AM onward.",
                        [
                            (30, "Kothapet"),
                            (31, "Chaitanyapuri"),
                            (32, "Gaddiannaram"),
                            (33, "Saroornagar"),
                            (34, "Doctors Colony"),
                            (35, "RK Puram"),
                            (36, "NTR Nagar"),
                        ],
                    ),
                    (
                        "L.B. Nagar",
                        "Busy commercial hub; split into two trips; afternoon + evening runs.",
                        [
                            (37, "Lingojiguda"),
                            (38, "Champapet"),
                            (39, "Kharmanghat"),
                            (40, "Bairamalguda"),
                            (41, "Hastinapuram"),
                        ],
                    ),
                    (
                        "Hayathnagar",
                        "Fringe run up to Vanasthalipuram; one round trip per day; return via LB Nagar ring.",
                        [
                            (42, "BN Reddy Nagar"),
                            (43, "Vanasthalipuram"),
                            (44, "Chintalkunta"),
                            (48, "High Court Colony"),
                            (49, "Sahebnagar"),
                            (50, "Hayathnagar"),
                        ],
                    ),
                ],
            },
        ],
    },
    "R4": {
        "name": "South-West (SW) Route",
        "zones": [
            {
                "zone": "Serilingampally Zone - South-West Belt",
                "circles": [
                    (
                        "Serilingampally",
                        "ORR tech belt (Gachibowli-Kondapur); peak 8-11 AM & 5-8 PM; deliver 11 AM-4 PM.",
                        [
                            (225, "Gachibowli"),
                            (226, "Nallagandla"),
                            (227, "Serilingampally"),
                            (228, "Masjid Banda"),
                            (229, "Sri Ram Nagar"),
                            (234, "Kondapur"),
                        ],
                    ),
                    (
                        "Madhapur",
                        "HITEC City core; parking premium; last-mile on 2W; schedule after 11 AM.",
                        [
                            (230, "Anjaiah Nagar"),
                            (231, "HITEC City"),
                            (232, "Madhapur"),
                            (233, "Izzath Nagar"),
                            (235, "Matrusri Nagar"),
                            (242, "Mayuri Nagar"),
                        ],
                    ),
                    (
                        "Narsingi",
                        "ORR new corridor; wide roads; low stop density; morning loop.",
                        [
                            (124, "Narsingi"),
                            (125, "Kokapet"),
                            (126, "Gandipet"),
                            (127, "Manikonda"),
                            (128, "Neknampur"),
                        ],
                    ),
                ],
            },
            {
                "zone": "Kukatpally Zone - South-West Wing",
                "circles": [
                    (
                        "Moosapet",
                        "Balanagar industrial + residential mix; truck-friendly mid-day; combine Kukatpally.",
                        [
                            (255, "Allapur"),
                            (256, "Moti Nagar"),
                            (257, "Moosapet"),
                            (258, "Prashanth Nagar"),
                            (259, "Balanagar"),
                        ],
                    ),
                ],
            },
            {
                "zone": "Secunderabad Zone - West Wing",
                "circles": [
                    (
                        "Ameerpet",
                        "Begumpet/Ameerpet commercial core; metered curbside drops; avoid 9-11 AM peak; combine Yousufguda/SR Nagar.",
                        [
                            (200, "Begumpet"),
                            (201, "Ameerpet"),
                            (202, "SR Nagar"),
                            (203, "BK Guda"),
                            (204, "Sanathnagar"),
                        ],
                    ),
                ],
            },
            {
                "zone": "Khairatabad Zone - West Wing",
                "circles": [
                    (
                        "Jubilee Hills",
                        "High-end gated societies; appointment-based; concierge delivery preferred.",
                        [
                            (215, "Jubilee Hills"),
                            (216, "Venkateshwara Colony"),
                            (221, "Banjara Hills"),
                            (222, "Film Nagar"),
                        ],
                    ),
                    (
                        "Yousufguda",
                        "Residential (Srinagar Colony, Yousufguda, Erragadda); combine Ameerpet/SR Nagar; mid-morning.",
                        [
                            (205, "Erragadda"),
                            (206, "Vengal Rao Nagar"),
                            (207, "Srinagar Colony"),
                            (208, "Yousufguda"),
                            (209, "AG Colony"),
                        ],
                    ),
                    (
                        "Borabanda",
                        "Residential (Borabanda, Karmika Nagar); narrow roads; 2W last mile; combine Yousufguda.",
                        [
                            (210, "Krishna Nagar"),
                            (211, "Rahamath Nagar"),
                            (212, "Karmika Nagar"),
                            (213, "Rajeev Nagar"),
                            (214, "Borabanda"),
                        ],
                    ),
                ],
            },
            {
                "zone": "Charminar Zone - Old City (South-West)",
                "circles": [
                    (
                        "Falaknuma",
                        "Old City core; narrow lanes; 2W only; 6-11 AM window.",
                        [
                            (104, "Shah Ali Banda"),
                            (105, "Falaknuma"),
                            (106, "Jahanuma"),
                            (107, "Nawab Saheb Kunta"),
                        ],
                    ),
                    (
                        "Bahadurpura",
                        "Dense old city; plan around Friday prayer times; 6-10 AM deliveries.",
                        [
                            (103, "Doodh Bowli"),
                            (108, "Teegal Kunta"),
                            (109, "Chandu Lal Baradari"),
                            (110, "Ramnasthpura"),
                            (111, "Kishanbagh"),
                        ],
                    ),
                    (
                        "Jangammet",
                        "Old City fringe; combine Yakutpura; early slots only.",
                        [
                            (73, "Riyasat Nagar"),
                            (74, "Lalitha Bagh"),
                            (75, "Jangammet"),
                            (76, "Phool Bagh"),
                            (77, "Quadri Chaman"),
                        ],
                    ),
                    (
                        "Charminar",
                        "Heritage core with vehicle curfews; offload at feeder point; hand-cart final mile.",
                        [
                            (97, "Purani Haveli"),
                            (98, "Pathergatti"),
                            (99, "Hari Bowli"),
                            (100, "Qazipura"),
                            (101, "Ghansi Bazar"),
                            (102, "Purana Pul"),
                        ],
                    ),
                ],
            },
            {
                "zone": "Golconda/Rajendranagar Zone - South-West Belt",
                "circles": [
                    (
                        "Goshamahal",
                        "Central market belt (Begum Bazar, Jambagh, Goshamahal); dense; early morning or post-8 PM drops.",
                        [
                            (148, "Dattatreya Nagar"),
                            (149, "Manghalhat"),
                            (150, "Goshamahal"),
                            (151, "Begum Bazar"),
                            (152, "Jambagh"),
                            (153, "Exhibition Grounds"),
                        ],
                    ),
                    (
                        "Karwan",
                        "Central-west (Karwan, Langar Houz, Gudimalkapur, Tappachabutra); market rush; combine Mehdipatnam.",
                        [
                            (134, "Langar Houz"),
                            (135, "Gudimalkapur"),
                            (136, "Karwan"),
                            (137, "Tappachabutra"),
                            (138, "Ziaguda"),
                        ],
                    ),
                    (
                        "Golconda",
                        "West-south belt (Toli Chowki, Golconda fort, Shaikpet, Nanalnagar); narrow fort lanes; morning window.",
                        [
                            (129, "Nizam Colony"),
                            (130, "Nanalnagar"),
                            (131, "Tolichowki"),
                            (132, "Golconda"),
                            (133, "Ibrahimbagh"),
                            (223, "Shaikpet"),
                            (224, "OU Colony"),
                        ],
                    ),
                    (
                        "Attapur",
                        "South-west fringe (Attapur, Katedan, Mailardevpally); combine Rajendranagar; mid-day.",
                        [
                            (112, "Attapur"),
                            (113, "Hyderguda"),
                            (114, "Suleman Nagar"),
                            (115, "Shastripuram"),
                            (116, "Katedan"),
                            (117, "Mailardevpally"),
                        ],
                    ),
                    (
                        "Rajendranagar",
                        "South fringe (Rajendra Nagar, Bandlaguda Jagir, Kismatpur); combine Attapur; avoid college rush 9-10 AM.",
                        [
                            (120, "Rajendra Nagar"),
                            (121, "Bandlaguda Jagir"),
                            (122, "Kismatpur"),
                            (123, "Hydershahkote"),
                        ],
                    ),
                ],
            },
            {
                "zone": "Khairatabad Zone - Central-South Belt",
                "circles": [
                    (
                        "Khairatabad",
                        "Central govt/corporate corridor; 10 AM-5 PM delivery windows.",
                        [
                            (146, "Red Hills"),
                            (154, "Gunfoundry"),
                            (217, "Irrum Manzil"),
                            (218, "Somajiguda"),
                            (219, "Khairatabad"),
                            (220, "Himayathnagar"),
                        ],
                    ),
                    (
                        "Mehdipatnam",
                        "Busy market + residential; after 11 AM; small vehicles.",
                        [
                            (139, "Asif Nagar"),
                            (140, "Padmanabha Nagar"),
                            (141, "Mehdipatnam"),
                            (142, "Syed Nagar"),
                        ],
                    ),
                    (
                        "Masab Tank",
                        "Dense mixed belt; 2W friendly; morning + evening peaks.",
                        [
                            (143, "Vijayanagar Colony"),
                            (144, "Ahmed Nagar"),
                            (145, "Shanti Nagar"),
                            (147, "Mallepally"),
                        ],
                    ),
                ],
            },
        ],
    },
    "R5": {
        "name": "Outside GHMC",
        "zones": [
            {
                "zone": "Outside GHMC (non-municipal)",
                "circles": [
                    (
                        "Sangareddy District",
                        "Long-haul via NH161; consolidate to truck load; weekly dispatch from Bowenpally hub.",
                        [
                            (None, "Sangareddy, Zaheerabad, Sadasivpet, Narayankhed, Patancheru fringe"),
                        ],
                    ),
                    (
                        "Medchal-Malkajgiri Rural",
                        "NH44 north fringe; Medchal check-post documentation; truck only; weekly run.",
                        [
                            (None, "Medchal, Shamirpet, Rampally, Baswapur, Kompally fringe"),
                        ],
                    ),
                    (
                        "Shamshabad (Rangareddy)",
                        "RGIA / NH44 south; airport-adjacent; truck; budget tolls; weekly run.",
                        [
                            (None, "Shamshabad, RGIA Airport, Tukkuguda, Bahadurguda"),
                        ],
                    ),
                    (
                        "Adibatla (2026-merged)",
                        "Airport-industrial belt (Adibatla, Kongara Kalan); IDA parks; truck; combine Shamshabad run.",
                        [
                            (53, "Thorrur"),
                            (54, "Kongara Kalan"),
                            (55, "Adibatla"),
                            (56, "Turkayamjal"),
                        ],
                    ),
                    (
                        "Badangpet (2026-merged)",
                        "Airport-east belt (Meerpet, Badangpet, Balapur); truck/LCV; combine Adibatla.",
                        [
                            (57, "Nadargul"),
                            (58, "Prashanthi Hills"),
                            (59, "Jillelaguda"),
                            (60, "Meerpet"),
                            (61, "Badangpet"),
                            (62, "Balapur"),
                        ],
                    ),
                    (
                        "Jalpally (2026-merged)",
                        "Airport-north belt (Shaheen Nagar, Pahadi Shareef); combine Badangpet; morning.",
                        [
                            (63, "Shaheen Nagar"),
                            (64, "Pahadi Shareef"),
                            (65, "Jalpally"),
                        ],
                    ),
                    (
                        "Ghatkesar Belt",
                        "NH163 east fringe; IDA Ghatkesar industrial; truck; weekly consolidation.",
                        [
                            (None, "Ghatkesar, Nagaram, Pocharam, Edulabad, Ibrahimpatnam"),
                        ],
                    ),
                    (
                        "Yadadri-Bhuvanagiri District",
                        "NH365 long haul; overnight trip; Yadagirigutta temple-town seasonality; truck only.",
                        [
                            (None, "Bhongir (Bhuvanagiri), Yadagirigutta, Alair, Choutuppal, Bibinagar"),
                        ],
                    ),
                ],
            }
        ],
    },
}

# Circle centroids (lat, lng) for the interactive map. Geocoded from
# Nominatim (OpenStreetMap) against the circle name + Hyderabad, India.
CIRCLE_COORDS = {
    "Bowenpally": [17.475537, 78.479228],
    "Alwal": [17.502229, 78.508858],
    "Jeedimetla": [17.519687, 78.446888],
    "Gajularamaram": [17.527176, 78.420008],
    "Nizampet": [17.497127, 78.376883],
    "Chintal": [17.502176, 78.440609],
    "Kompally": [17.535487, 78.509698],
    "Dundigal": [17.50621, 78.505108],
    "Kapra": [17.484636, 78.56101],
    "Uppal": [17.402509, 78.561256],
    "Boduppal": [17.398841, 78.536622],
    "Nacharam": [17.428494, 78.55281],
    "Malkajgiri": [17.451176, 78.5369],
    "Moula Ali": [17.46171, 78.55714],
    "Keesara": [17.481583, 78.592682],
    "Tarnaka": [17.428548, 78.537943],
    "Mettuguda": [17.435504, 78.519557],
    "Serilingampally": [17.466717, 78.340421],
    "Madhapur": [17.440892, 78.39163],
    "Miyapur": [17.498161, 78.356763],
    "Narsingi": [17.387417, 78.356624],
    "Patancheru": [17.528609, 78.267425],
    "Ameenpur": [17.523691, 78.33173],
    "Kukatpally": [17.493084, 78.405441],
    "Moosapet": [17.468531, 78.42067],
    "Allwyn Colony": [17.492035, 78.349953],
    "Ameerpet": [17.437501, 78.448251],
    "Jubilee Hills": [17.430836, 78.410288],
    "Yousufguda": [17.43875, 78.427987],
    "Borabanda": [17.459069, 78.407866],
    "Falaknuma": [17.33266, 78.475198],
    "Bahadurpura": [17.357067, 78.454542],
    "Chandrayangutta": [17.324696, 78.481356],
    "Yakutpura": [17.358628, 78.485805],
    "Jangammet": [17.336251, 78.474404],
    "Santoshnagar": [17.346719, 78.508195],
    "Charminar": [17.361602, 78.474642],
    "Malakpet": [17.373671, 78.499648],
    "Moosarambagh": [17.374353, 78.516084],
    "Goshamahal": [17.380576, 78.468846],
    "Karwan": [17.376013, 78.433189],
    "Golconda": [17.387329, 78.405734],
    "Attapur": [17.367224, 78.430728],
    "Rajendranagar": [17.334621, 78.40868],
    "Khairatabad": [17.412974, 78.461058],
    "Mehdipatnam": [17.394263, 78.434251],
    "Masab Tank": [17.402962, 78.450754],
    "Kavadiguda": [17.422702, 78.49177],
    "Musheerabad": [17.419142, 78.498573],
    "Amberpet": [17.386178, 78.511471],
    "Nagole": [17.377531, 78.560123],
    "Saroornagar": [17.361166, 78.538744],
    "L.B. Nagar": [17.349807, 78.547888],
    "Hayathnagar": [17.328115, 78.60454],
    "Sangareddy District": [17.528026, 78.267025],
    "Medchal-Malkajgiri Rural": [17.633993, 78.484315],
    "Shamshabad (Rangareddy)": [17.257207, 78.345104],
    "Adibatla (2026-merged)": [17.230899, 78.5559],
    "Badangpet (2026-merged)": [17.338347, 78.522213],
    "Jalpally (2026-merged)": [17.306154, 78.473859],
    "Ghatkesar Belt": [17.451084, 78.684302],
    "Yadadri-Bhuvanagiri District": [17.517279, 78.886338],
}

COLUMNS = [
    "Route Code",
    "Route Name",
    "GHMC Zone",
    "GHMC Circle Name",
    "GHMC Ward Number & Name",
    "Logistics Sorting Note",
    "Circle Lat",
    "Circle Lng",
]

# --------------------------------------------------------------------------- #
# 2. Locality layer                                                           #
# --------------------------------------------------------------------------- #
# Two structures feed locality rows:
#   LOCALITY_GROUPS : (route, circle) -> (shared_note, [well-known localities])
#                     bulk neighbourhoods / sub-areas under their circle.
#   LOCALITIES      : name -> (route, circle, note)  custom-note localities
#                     plus the exact spellings used in the live outward sheet
#                     (including typos/alt spellings) so raw data joins cleanly.
LOCALITY_GROUPS = {
    # ---- R1 North-East (NE) -------------------------------------------------
    ("R1", "Bowenpally"): (
        "Hub-adjacent (Bowenpally); first sort batch; dispatch before 08:30 AM.",
        ["Old Bowenpally", "Hasmathpet", "Bowenpally Crossroads", "Bolarum"],
    ),
    ("R1", "Alwal"): (
        "North-east wing on NH44 flank; batch with Quthbullapur belt; morning dispatch.",
        ["Old Alwal", "Rathifile", "Temple Alwal", "Cantonment Alwal"],
    ),
    ("R1", "Kompally"): (
        "Quthbullapur north-east; gated layouts; morning dispatch.",
        ["Kompally (Ecopark)", "Doolapally", "Almasguda", "Bachupally (north)"],
    ),
    ("R1", "Dundigal"): (
        "Quthbullapur far north-east; drone/air-force flank; morning only.",
        ["Dundigal", "Bowrampet", "Bahadurpally"],
    ),
    ("R1", "Kapra"): (
        "ECIL & A.S. Rao Nagar belt; security gate protocols; 10 AM-3 PM.",
        ["ECIL X Roads", "A.S. Rao Nagar", "Kushaiguda", "Cherlapally",
         "Vampuguda", "Dr. A.S. Rao Nagar"],
    ),
    ("R1", "Keesara"): (
        "North-east fringe via Nagaram road; one morning loop.",
        ["Dammaiguda", "Yapral", "Nagaram", "Chandrapuri Colony", "Keesara Gutta"],
    ),
    # ---- R2 North-West (NW) -------------------------------------------------
    ("R2", "Jeedimetla"): (
        "IDA Jeedimetla industrial belt; truck dock 10 AM-4 PM; gate passes required.",
        ["IDA Jeedimetla", "Pet Basheerabad", "Suraram Colony"],
    ),
    ("R2", "Gajularamaram"): (
        "Pharma/industrial corridor; gate-based drop-offs; combine Alwal loop.",
        ["IDA Gajularamaram", "Mahadevpuram", "Shapur Nagar"],
    ),
    ("R2", "Nizampet"): (
        "Residential north-west wing (Bachupally); off-peak after 10 AM.",
        ["Bachupally", "Pragathi Nagar", "Kukatpally Housing Board fringe"],
    ),
    ("R2", "Chintal"): (
        "Quthbullapur industrial north-west; truck-friendly; combine Jeedimetla.",
        ["Chintal", "Jagathgiri Gutta", "Giri Nagar"],
    ),
    ("R2", "Kukatpally"): (
        "KPHB high-density residential; small vehicles; 6-10 AM window.",
        ["Old Kukatpally", "Kukatpally Housing Board", "KPHB Colony",
         "Balaji Nagar", "Vasanth Nagar", "Kaithalapur", "Gayatri Nagar"],
    ),
    ("R2", "Miyapur"): (
        "Chandanagar/Hafeezpet/Miyapur on NH65; combine Kukatpally south; mid-day.",
        ["Chandanagar", "Madeenaguda", "Deepthisri Nagar", "Old Hafeezpet",
         "Maktha Mahabubpet"],
    ),
    ("R2", "Allwyn Colony"): (
        "Kukatpally south residential; narrow lanes; 2W last mile.",
        ["Hyder Nagar", "Shamshiguda", "Bhagya Nagar Colony", "Vivekananda Nagar Colony"],
    ),
    ("R2", "Patancheru"): (
        "West fringe on NH65; industrial belt; truck-friendly; plan toll.",
        ["ICRISAT", "IDA Patancheru", "Tellapur", "Muthangi", "JP Colony"],
    ),
    ("R2", "Ameenpur"): (
        "West fringe (RC Puram, Beeramguda); NH65 corridor; mid-day.",
        ["RC Puram", "Beeramguda", "Bharathi Nagar", "Bollaram"],
    ),
    # ---- R3 South-East (SE) -------------------------------------------------
    ("R3", "Uppal"): (
        "IT/industrial mix; IDA Uppal truck gates till 4 PM; combine Boduppal.",
        ["IDA Uppal", "Habsiguda", "Ramanthapur", "Chilkanagar", "Uppal Depot"],
    ),
    ("R3", "Boduppal"): (
        "Fast-growing corridor; wide roads; multiple quick stops.",
        ["Medipally", "Peerzadiguda", "Chengicherla", "Boduppal X Road"],
    ),
    ("R3", "Nacharam"): (
        "IDA Nacharam industrial; ORR feeder restrictions 6-9 PM; mid-day runs.",
        ["IDA Nacharam", "Mallapur", "HMT Nagar", "Shakthi Sai Nagar"],
    ),
    ("R3", "Malkajgiri"): (
        "Railway colony + gated societies; combine Kapra/Nacharam loop; early start.",
        ["Safilguda", "Old Malkajgiri", "Malkajgiri Railway Colony", "Anandbagh"],
    ),
    ("R3", "Moula Ali"): (
        "East-central (Neredmet, Moula Ali); combine Malkajgiri/Sainikpuri loop.",
        ["Neredmet", "Sainikpuri", "Vinayak Nagar", "Kakatiya Nagar", "Balram Nagar"],
    ),
    ("R3", "Tarnaka"): (
        "Defence-research corridor; gate-entry paperwork; 15-min dwell per stop.",
        ["Osmania University", "Lalapet", "AOC Camp", "Golconda X Roads (Tarnaka)"],
    ),
    ("R3", "Mettuguda"): (
        "Railway-colony lanes; narrow roads; 2W/LCV only; morning slots.",
        ["Lalaguda", "Old Tarnaka", "Metro JBS flank"],
    ),
    ("R3", "Kavadiguda"): (
        "Secunderabad core + cantonment belt; avoid 5-7 PM rush; LCV/2W.",
        [
            "Trimulgherry", "Marredpally", "Rasoolpura", "Regimental Bazaar",
            "Parsigutta", "North Secunderabad", "Rani Gunj", "General Bazaar",
            "Mahatma Gandhi Road", "JBS Paradise", "Patny", "King Koti",
            "Tadbund", "Sangeet Nagar", "Bhoiguda",
        ],
    ),
    ("R3", "Amberpet"): (
        "OU/university belt + market; schedule after 11 AM; combine Kavadiguda/Musheerabad belt.",
        ["Vidyanagar", "OU Campus", "Kachiguda (railway)", "Chikoti Gardens"],
    ),
    ("R3", "Chandrayangutta"): (
        "Barkas belt; steep lanes; small vehicles; morning window.",
        ["Barkas", "Kanchanbagh", "Bandlaguda", "Noori Nagar"],
    ),
    ("R3", "Yakutpura"): (
        "Charminar market spill; pedestrian-heavy; early morning drops.",
        ["Dabeerpura", "Rein Bazar", "Madannapet", "Gowlipura", "Talab Chanchalam"],
    ),
    ("R3", "Santoshnagar"): (
        "Mixed density mid-commercial; schedule 10 AM-1 PM.",
        ["Talab Katta", "Bhanu Nagar", "IS Sadan", "Saraswati Nagar"],
    ),
    ("R3", "Malakpet"): (
        "Commercial mid-belt; combine Moosarambagh; post-11 AM.",
        ["Malakpet", "Chaderghat", "Saidabad", "Asmangadh", "Akberbagh", "Chawani"],
    ),
    ("R3", "Moosarambagh"): (
        "Ring-road access; quick drops; combine Malakpet.",
        ["Old Malakpet", "Kala Dera", "Azampura", "MCH Colony"],
    ),
    ("R3", "Saroornagar"): (
        "High-density residential/market; avoid 9-11 AM; schedule 11 AM onward.",
        ["Dilsukhnagar", "Kothapet", "Chaitanyapuri", "Gaddiannaram",
         "P&T Colony", "Saroornagar Lake", "Laxminagar", "Alkapuri Colony"],
    ),
    ("R3", "L.B. Nagar"): (
        "Busy commercial hub; split into two trips; afternoon + evening runs.",
        ["Lingojiguda", "Champapet", "Kharmanghat", "Bairamalguda",
         "Hastinapuram", "Buddha Nagar"],
    ),
    ("R3", "Nagole"): (
        "South-east gateway; start after 9 AM; combine Saroornagar/LB Nagar.",
        ["Mansoorabad", "Kuntloor", "Pedda Amberpet", "GSI Colony"],
    ),
    ("R3", "Hayathnagar"): (
        "Fringe up to Vanasthalipuram; one round trip/day; return via LB Nagar ring.",
        ["Vanasthalipuram", "Sahebnagar", "BN Reddy Nagar", "Nagarjuna Sagar Road stretch"],
    ),
    # ---- R4 South-West (SW) -------------------------------------------------
    ("R4", "Serilingampally"): (
        "ORR tech belt (Gachibowli-Kondapur); deliver 11 AM-4 PM.",
        ["Raidurg", "Nanakramguda", "Gopanpally", "Kothaguda", "Nallagandla",
         "Financial District", "Inorbit Mall", "The HUB", "Masjid Banda"],
    ),
    ("R4", "Madhapur"): (
        "HITEC City core; parking premium; last-mile on 2W; after 11 AM.",
        ["Cyber Towers", "Durgam Cheruvu", "Jubilee Enclave", "Izzath Nagar",
         "Matrusri Nagar", "Mayuri Nagar", "Mindspace"],
    ),
    ("R4", "Narsingi"): (
        "ORR new corridor; wide roads; low stop density; morning loop.",
        ["Kokapet", "Manikonda", "Gandipet", "Neknampur", "Narsingi (ORR)"],
    ),
    ("R4", "Moosapet"): (
        "Balanagar industrial + residential; truck-friendly mid-day.",
        ["IDA Balanagar", "Ferozguda", "Moti Nagar", "Prashanth Nagar", "Allapur"],
    ),
    ("R4", "Ameerpet"): (
        "Ameerpet/Begumpet commercial core; metered curbside drops; avoid 9-11 AM peak.",
        ["Balkampet", "West Marredpally", "Divya Nagar", "Sai Baba Temple Road"],
    ),
    ("R4", "Jubilee Hills"): (
        "High-end gated societies; appointment-based; concierge delivery preferred.",
        ["Banjara Hills", "Film Nagar", "Road No. 12", "Venkateshwara Colony"],
    ),
    ("R4", "Yousufguda"): (
        "Residential (Srinagar Colony); combine Ameerpet/SR Nagar; mid-morning.",
        ["Srinagar Colony", "Erragadda", "Vengal Rao Nagar", "AG Colony"],
    ),
    ("R4", "Borabanda"): (
        "Residential; narrow roads; 2W last mile; combine Yousufguda.",
        ["Karmika Nagar", "Rajeev Nagar", "Krishna Nagar", "Rahamath Nagar"],
    ),
    ("R4", "Charminar"): (
        "Heritage core with vehicle curfews; offload at feeder point; hand-cart final mile.",
        ["Afzalgunj", "Moghalpura", "Aliabad", "Darulshifa", "Hussaini Alam",
         "Purani Haveli", "Pathergatti"],
    ),
    ("R4", "Falaknuma"): (
        "Old City core; narrow lanes; 2W only; 6-11 AM window.",
        ["Falaknuma Palace Road", "Shah Ali Banda", "Jahanuma", "Nawab Saheb Kunta"],
    ),
    ("R4", "Bahadurpura"): (
        "Dense old city; plan around Friday prayer; 6-10 AM deliveries.",
        ["Bahadurpura", "Doodh Bowli", "Kishanbagh", "Teegal Kunta",
         "Chandu Lal Baradari", "Ramnasthpura"],
    ),
    ("R4", "Goshamahal"): (
        "Central market belt; dense; early morning or post-8 PM drops.",
        ["Begum Bazar", "Jambagh", "Dattatreya Nagar", "Exhibition Grounds", "Koti"],
    ),
    ("R4", "Karwan"): (
        "Central-west market belt; combine Mehdipatnam.",
        ["Langar Houz", "Gudimalkapur", "Tappachabutra", "Ziaguda"],
    ),
    ("R4", "Golconda"): (
        "West-south belt; narrow fort lanes; morning window.",
        ["Tolichowki", "Shaikpet", "Nanalnagar", "Nizam Colony", "Ibrahimbagh",
         "Golconda Fort", "OU Colony"],
    ),
    ("R4", "Mehdipatnam"): (
        "Busy market + residential; after 11 AM; small vehicles.",
        ["Asif Nagar", "Syed Nagar", "Padmanabha Nagar", "Gudimalkapur fringe"],
    ),
    ("R4", "Masab Tank"): (
        "Dense mixed belt; 2W friendly; morning + evening peaks.",
        ["Public Gardens", "Saifabad", "Lakdikapul", "Vijayanagar Colony", "Ahmed Nagar"],
    ),
    ("R4", "Khairatabad"): (
        "Central govt/corporate corridor; 10 AM-5 PM windows.",
        ["Greenlands", "Gunfoundry", "Red Hills", "Somajiguda", "Irrum Manzil"],
    ),
    ("R4", "Attapur"): (
        "South-west fringe; combine Rajendranagar; mid-day.",
        ["Katedan", "Mailardevpally", "Hyderguda", "Suleman Nagar", "Shastripuram"],
    ),
    ("R4", "Rajendranagar"): (
        "South fringe; avoid college rush 9-10 AM; combine Attapur.",
        ["Bandlaguda Jagir", "Kismatpur", "Hydershahkote"],
    ),
    # ---- R5 Outside GHMC ----------------------------------------------------
    ("R5", "Medchal-Malkajgiri Rural"): (
        "NH44 north fringe; Medchal check-post docs; truck only; weekly run.",
        ["Medchal", "Shamirpet", "Rampally", "Baswapur", "Thumkunta"],
    ),
    ("R5", "Ghatkesar Belt"): (
        "NH163 east fringe; IDA Ghatkesar industrial; truck; weekly consolidation.",
        ["Ghatkesar", "Nagaram", "Pocharam", "Edulabad", "Ibrahimpatnam"],
    ),
}

LOCALITIES = {
    # custom-note localities (key delivery points) -----------------------------
    "Secunderabad": ("R3", "Kavadiguda",
        "Secunderabad city centre (Parade Grounds, Clock Tower); avoid 5-7 PM rush; LCV/2W."),
    "Karkhana": ("R1", "Alwal",
        "Army supply area; gate protocols; combine Alwal loop."),
    "Dilsukhnagar": ("R3", "Saroornagar",
        "Major commercial hub on LB Nagar ring; metro corridor; two trips; post-10 AM."),
    "ECIL": ("R1", "Kapra",
        "ECIL township; security check at gates; 10 AM-4 PM."),
    "LB Nagar": ("R3", "L.B. Nagar",
        "L.B. Nagar commercial hub; split into two trips; afternoon + evening runs."),
    "Financial District": ("R4", "Serilingampally",
        "ORR/Gachibowli offices; 11 AM-4 PM window; corporate receptions."),
    "Gowliguda": ("R4", "Charminar",
        "Old City belt near Koti/Chaderghat; narrow lanes; 2W only; morning window."),
    "Koti": ("R4", "Goshamahal",
        "Commercial + textile market; pedestrian-heavy; early drops; hand-cart final mile."),
    "Abids": ("R4", "Khairatabad",
        "Commercial centre near Gunfoundry; metered parking; post-11 AM."),
    "Nampally": ("R4", "Masab Tank",
        "Railway station + hotel belt; early morning + evening peaks."),
    "Punjagutta": ("R4", "Khairatabad",
        "Commercial corridor (Banjara Hills/Greenlands); combine Somajiguda; mid-day."),
    "Chaderghat": ("R3", "Malakpet",
        "Musi-bridge belt between old city & LB Nagar; combine Malakpet."),
    "Sainikpuri": ("R3", "Moula Ali",
        "Defence colony; gated; appointments preferred; combine Neredmet."),
    "Kompally": ("R1", "Kompally",
        "Quthbullapur north-east; gated layouts; morning dispatch with Quthbullapur belt."),
    "Suchitra": ("R2", "Jeedimetla",
        "Quthbullapur residential circle on the ring road (north of Jeedimetla); gated layouts; combine Jeedimetla return loop."),
    # live-sheet spelling variants (typos / alt spellings) --------------------
    "Gcachibowli": ("R4", "Serilingampally",
        "Alt. spelling of Gachibowli; ORR tech belt; deliver 11 AM-4 PM."),
    "Kkatpally": ("R2", "Kukatpally",
        "Alt. spelling of Kukatpally; KPHB high-density; 6-10 AM window."),
    "Malkpet": ("R3", "Malakpet",
        "Alt. spelling of Malakpet; commercial mid-belt; post-11 AM."),
    "Mehndipatnam": ("R4", "Mehdipatnam",
        "Alt. spelling of Mehdipatnam; market + residential; after 11 AM."),
    "Myapur": ("R2", "Miyapur",
        "Alt. spelling of Miyapur; NH65 corridor; mid-day."),
    "Hitech City": ("R4", "Madhapur",
        "Alt. spelling of HITEC City (ward 231); same routing as Madhapur."),
    "Himayatnagar": ("R4", "Khairatabad",
        "Alt. spelling of Himayathnagar (ward 220); central corporate; 10 AM-5 PM."),
    "Suchitra Circle": ("R2", "Jeedimetla",
        "Alt. spelling of Suchitra; Quthbullapur residential circle; combine Jeedimetla return loop."),
}

# --------------------------------------------------------------------------- #
# 3. Build flat row list                                                      #
# --------------------------------------------------------------------------- #
def _zone_for(route_code: str, circle_name: str) -> str:
    """Resolve the zone label for a route + circle (used by the locality layer)."""
    for zone in DATA[route_code]["zones"]:
        for cname, _note, _wards in zone["circles"]:
            if cname == circle_name:
                return zone["zone"]
    raise KeyError(f"Circle '{circle_name}' not found on route {route_code}")


def build_records() -> list[dict]:
    """Flatten DATA + LOCALITY_GROUPS + LOCALITIES into one dict per row."""
    records = []
    seen = set()  # (route, circle, name_lower) to avoid duplicates

    def add(route_code: str, zone: str, circle_name: str, ward_col: str,
            note: str, dedupe_key: str | None) -> None:
        if dedupe_key is not None:
            key = (route_code, circle_name, dedupe_key.lower())
            if key in seen:
                return
            seen.add(key)
        lat, lng = CIRCLE_COORDS.get(circle_name, (None, None))
        records.append(
            {
                "Route Code": route_code,
                "Route Name": DATA[route_code]["name"],
                "GHMC Zone": zone,
                "GHMC Circle Name": circle_name,
                "GHMC Ward Number & Name": ward_col,
                "Logistics Sorting Note": note,
                "Circle Lat": lat,
                "Circle Lng": lng,
            }
        )

    for route_code, route in DATA.items():
        for zone in route["zones"]:
            for circle_name, note, wards in zone["circles"]:
                for ward_no, ward_name in wards:
                    if ward_no is None:
                        ward_col = "- (non-GHMC): " + ward_name
                    else:
                        ward_col = f"{ward_no} - {ward_name}"
                    add(route_code, zone["zone"], circle_name, ward_col, note,
                        dedupe_key=ward_name)

    # bulk locality groups (shared note per circle)
    for (route_code, circle_name), (note, names) in LOCALITY_GROUPS.items():
        zone = _zone_for(route_code, circle_name)
        for name in names:
            add(route_code, zone, circle_name, f"- (locality): {name}", note,
                dedupe_key=name)

    # custom-note localities + live-sheet spellings
    for name, (route_code, circle_name, note) in LOCALITIES.items():
        zone = _zone_for(route_code, circle_name)
        add(route_code, zone, circle_name, f"- (locality): {name}", note,
            dedupe_key=name)

    return records


# --------------------------------------------------------------------------- #
# 3. Excel formatting helpers                                                 #
# --------------------------------------------------------------------------- #
NAVY = "1F3864"
WHITE = "FFFFFF"
ZEBRA = "F2F2F2"
GRID = "BFBFBF"
THIN = Side(style="thin", color=GRID)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
BODY_FONT = Font(name="Calibri", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor=ZEBRA)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _display_width(value) -> int:
    """Best-effort display width of a cell value (CJK-safe approximation)."""
    return max(len(line) for line in str(value).split("\n"))


def autofit_columns(ws, df, cap: int = 70) -> None:
    """Size each column to the longest visible value, with a hard cap."""
    for idx, col in enumerate(df.columns, start=1):
        header_len = _display_width(col)
        data_len = max((_display_width(v) for v in df[col].tolist()), default=0)
        width = min(cap, max(header_len, data_len) * 1.08 + 3)
        ws.column_dimensions[get_column_letter(idx)].width = width


def style_workbook(ws, df) -> None:
    """Apply header, zebra, gridlines, widths, freeze panes and auto-filter."""
    n_rows = df.shape[0] + 1
    n_cols = df.shape[1]

    # Header row ------------------------------------------------------------
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BOX
    ws.row_dimensions[1].height = 30

    # Body rows with zebra striping + explicit gridlines ---------------------
    for row in range(2, n_rows + 1):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = BOX
            if row % 2 == 0:
                cell.fill = ZEBRA_FILL

    # Gridlines on screen and in print ----------------------------------------
    ws.sheet_view.showGridLines = True
    ws.print_options.gridLines = True

    # Freeze header, auto-filter over full used range --------------------------
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows}"

    # Print layout --------------------------------------------------------------
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:1"
    ws.sheet_view.zoomScale = 90

    autofit_columns(ws, df)


# --------------------------------------------------------------------------- #
# 4. JSON output                                                              #
# --------------------------------------------------------------------------- #
def write_json(records: list[dict], out_path: Path) -> None:
    payload = _build_payload(records)
    out_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def _build_payload(records: list[dict]) -> dict:
    return {
        "meta": {
            "title": "SkyLimit Outward Delivery Routes - GHMC Administrative Map",
            "hub": "Bowenpally",
            "source": "GHMC 2026 zones/circles/wards (G.O.Ms.No.292, 24-12-2025)",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "row_count": len(records),
            "locality_count": sum(
                str(r["GHMC Ward Number & Name"]).startswith("- (locality):")
                for r in records
            ),
        },
        "records": records,
    }


def write_js_bundle(records: list[dict], out_path: Path) -> None:
    """Emit the same data as a plain script (window.GHMC_ROUTE_MAP = ...).

    The dashboard loads this via a <script> tag so routing + the interactive
    map work even when the page is opened from the file system, where
    fetch() to a local .json file is blocked by the browser's CORS policy.
    """
    payload = _build_payload(records)
    js = (
        "/* Auto-generated from ghmc-routing/output/ghmc_route_map.json by\n"
        "   generate_ghmc_route_map.py. Bundled as a plain script so the route\n"
        "   map works even when the app is opened from the file system (where\n"
        "   fetch() to local JSON is blocked by CORS).\n"
        "   Regenerate with:  python ghmc-routing/generate_ghmc_route_map.py\n"
        "*/\n"
        "window.GHMC_ROUTE_MAP = "
        + json.dumps(payload, ensure_ascii=False)
        + ";\n"
    )
    out_path.write_text(js, encoding="utf-8")


# --------------------------------------------------------------------------- #
# 5. Coverage checker                                                          #
# --------------------------------------------------------------------------- #
def build_index(records: list[dict]) -> dict[str, dict]:
    """Lowercased name -> record, so any location can be looked up fast."""
    index: dict[str, dict] = {}
    for rec in records:
        name = str(rec["GHMC Ward Number & Name"])
        if name.startswith("- "):
            name = name.split(":", 1)[1].strip()
        index.setdefault(name.lower(), rec)
    return index


def _match_kind(name: str) -> str:
    if name.startswith("- (non-GHMC)"):
        return "non-GHMC area"
    if name.startswith("- (locality):"):
        return "locality"
    return "ward"


def resolve(location: str, index: dict[str, dict]) -> tuple[dict | None, str]:
    """Return (record, status) where status is one of exact|fuzzy|unmapped."""
    key = location.strip().lower()
    if not key:
        return None, "unmapped"
    if key in index:
        return index[key], "exact"
    best = None
    for name, rec in index.items():
        if key in name or name in key:
            if best is None or len(name) < len(best[0]):
                best = (name, rec)
    if best is not None:
        return best[1], "fuzzy"
    return None, "unmapped"


def read_locations(path: Path) -> list[str]:
    """One location per line (.txt) or a `Location` column (.csv/.tsv)."""
    text = path.read_text(encoding="utf-8", errors="replace")
    if path.suffix.lower() in {".csv", ".tsv"}:
        lines = [ln for ln in text.splitlines() if ln.strip()]
        header = lines[0].strip().lower()
        try:
            delim = "\t" if "\t" in header else ","
            cols = [c.strip().strip('"') for c in header.split(delim)]
            col = next(c for c in cols if "location" in c)
        except StopIteration:
            col = 0
        out = []
        for ln in lines[1:]:
            cells = [c.strip().strip('"') for c in ln.split(delim)]
            if len(cells) > col:
                out.append(cells[col])
        return out
    return [ln.strip() for ln in text.splitlines() if ln.strip()]


def check_coverage(path: Path) -> None:
    records = build_records()
    index = build_index(records)
    locations = read_locations(path)

    exact = fuzzy = unmapped = 0
    print(f"Checking {len(locations)} location(s) against "
          f"{len(records)} records ({len(index)} unique names)\n")
    for loc in locations:
        rec, status = resolve(loc, index)
        if rec is None:
            unmapped += 1
            print(f"  [UNMAPPED ] {loc}")
            continue
        kind = _match_kind(str(rec["GHMC Ward Number & Name"]))
        if status == "exact":
            exact += 1
        else:
            fuzzy += 1
            print(f"  [~fuzzy   ] {loc}  -> {rec['GHMC Ward Number & Name']}")
        print(f"    -> {rec['Route Code']} | {rec['GHMC Zone']} | "
              f"{rec['GHMC Circle Name']} ({kind})")
    print(f"\nResult: {exact} exact, {fuzzy} fuzzy, {unmapped} unmapped "
          f"of {len(locations)}")


# --------------------------------------------------------------------------- #
# 6. Main                                                                     #
# --------------------------------------------------------------------------- #
def main() -> None:
    parser = argparse.ArgumentParser(description="Build GHMC route-map xlsx + json")
    parser.add_argument(
        "--output-dir",
        default=str(Path(__file__).resolve().parent / "output"),
        help="Directory for the generated files",
    )
    parser.add_argument(
        "--check",
        metavar="FILE",
        help="Instead of generating, verify a list of locations "
             "(one per line, or CSV/TSV with a Location column)",
    )
    args = parser.parse_args()

    if args.check:
        check_coverage(Path(args.check))
        return

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    records = build_records()
    df = pd.DataFrame.from_records(records, columns=COLUMNS)

    # Excel ------------------------------------------------------------------
    excel_path = out_dir / "ghmc_route_map.xlsx"
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="GHMC Route Map")
        workbook = writer.book
        style_workbook(workbook["GHMC Route Map"], df)
    # Note: pandas 3.x writes on close; reopen is not needed, styling applied
    # through the live workbook object above before save.

    # JSON -------------------------------------------------------------------
    json_path = out_dir / "ghmc_route_map.json"
    write_json(records, json_path)

    # JS bundle (for the dashboard's offline / file:// mode) ------------------
    # Lives next to index.html at the repo root, not inside --output-dir.
    repo_root = Path(__file__).resolve().parent.parent
    js_path = repo_root / "ghmc-route-map.js"
    write_js_bundle(records, js_path)

    print(f"Wrote {excel_path}  ({df.shape[0]} rows x {df.shape[1]} cols)")
    print(f"Wrote {json_path}")
    print(f"Wrote {js_path}")


if __name__ == "__main__":
    main()
