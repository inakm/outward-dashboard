"""Fill routes-planning.xlsx (Routes sheet) from the GHMC generator DATA plus
the customer master (Customers/ALL_Customers*.xlsx).

Top section = canonical circle overview per route column (R1-R4 inside GHMC,
R5 = outside-GHMC fringe, R7 = other routes / outside Hyderabad).
Bottom section = complete customer address ledger (no skips): every customer
mapped to its route + circle, with full address.

Idempotent: safe to re-run whenever DATA or the customer file changes.
"""

import json
import re
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "ghmc-routing"))
from generate_ghmc_route_map import DATA, LOCALITIES, LOCALITY_GROUPS  # noqa: E402

CUSTOMERS = REPO / "Customers" / "ALL_Customers 31st july.xlsx"
XLSX = REPO / "routes-planning.xlsx"

ROUTE_ORDER = ["R1", "R2", "R3", "R4", "R5", "R7"]
COL_BY_ROUTE = {"R1": 2, "R2": 3, "R3": 4, "R4": 5, "R5": 6, "R7": 7}

NAVY = "1F3864"
WHITE = "FFFFFF"
ZEBRA = "F2F2F2"
GRID = "BFBFBF"
ZONE_FILL_HEX = "DCE6F1"
NOTE_FILL_HEX = "FFF2CC"
LEDGER_FILL_HEX = "D9E1F2"
REVIEW_FILL_HEX = "FBE5D6"

THIN = Side(style="thin", color=GRID)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
ZONE_FONT = Font(name="Calibri", size=11, bold=True, color=NAVY)
ZONE_FILL = PatternFill("solid", fgColor=ZONE_FILL_HEX)
BODY_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
ZEBRA_FILL = PatternFill("solid", fgColor=ZEBRA)
NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="7F6000")
NOTE_FILL = PatternFill("solid", fgColor=NOTE_FILL_HEX)
LEDGER_FILL = PatternFill("solid", fgColor=LEDGER_FILL_HEX)
REVIEW_FILL = PatternFill("solid", fgColor=REVIEW_FILL_HEX)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def norm(s):
    s = (s or "").lower()
    s = s.replace("\n", " ").replace("\r", " ").replace(",", " ").replace(".", " ")
    s = re.sub(r"[()/&#+'-]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def tight(s):
    return re.sub(r"\s+", "", norm(s))


# --------------------------------------------------------------------------- #
# Place index: known place -> (route, circle, zone)                            #
# --------------------------------------------------------------------------- #
known = {}


def add(name, route, circle, zone):
    t = tight(name)
    if len(t) >= 4:
        known.setdefault(t, (route, circle, zone))


for route, rd in DATA.items():
    for zone in rd["zones"]:
        for circle, _note, wards in zone["circles"]:
            add(circle, route, circle, zone["zone"])
            for _wno, wname in wards:
                add(wname, route, circle, zone["zone"])

for (route, circle), (_note, names) in LOCALITY_GROUPS.items():
    zone = None
    for z in DATA[route]["zones"]:
        if circle in [c[0] for c in z["circles"]]:
            zone = z["zone"]
    for name in names:
        add(name, route, circle, zone)

for name, (route, circle, _note) in LOCALITIES.items():
    zone = None
    for z in DATA[route]["zones"]:
        if circle in [c[0] for c in z["circles"]]:
            zone = z["zone"]
    add(name, route, circle, zone)

# Alias spellings / areas used in the customer file (incl. typos).
ALIASES = {
    # R1
    "ghorinagar": ("R1", "Bowenpally", "Secunderabad Zone - Hub Belt"),
    "ghori nagar": ("R1", "Bowenpally", "Secunderabad Zone - Hub Belt"),
    "kalyan nagar phase 1": ("R1", "Alwal", "Kukatpally/Quthbullapur Zone - North-East Wing"),
    "kalyan nagar": ("R1", "Alwal", "Kukatpally/Quthbullapur Zone - North-East Wing"),
    # R2
    "suchitra": ("R2", "Jeedimetla", "Kukatpally/Quthbullapur Zone - North-West Wings"),
    "suchitra circle": ("R2", "Jeedimetla", "Kukatpally/Quthbullapur Zone - North-West Wings"),
    "gundlapochampally": ("R2", "Nizampet", "Kukatpally/Quthbullapur Zone - North-West Wings"),
    "madinaguda": ("R2", "Miyapur", "Kukatpally Zone - North-West Wing"),
    "madhinaguda": ("R2", "Miyapur", "Kukatpally Zone - North-West Wing"),
    "mythri nagar": ("R2", "Miyapur", "Kukatpally Zone - North-West Wing"),
    # R3
    "vidyanagar": ("R3", "Amberpet", "Secunderabad Zone - South-East Wing"),
    "tilak nagar": ("R3", "Amberpet", "Secunderabad Zone - South-East Wing"),
    "tillak nagar": ("R3", "Amberpet", "Secunderabad Zone - South-East Wing"),
    "bayamma galli": ("R3", "Amberpet", "Secunderabad Zone - South-East Wing"),
    "narayanaguda": ("R3", "Amberpet", "Secunderabad Zone - South-East Wing"),
    "chikkadpally": ("R3", "Amberpet", "Secunderabad Zone - South-East Wing"),
    "gowra grand": ("R3", "Amberpet", "Secunderabad Zone - South-East Wing"),
    "dayanad nagar": ("R3", "Amberpet", "Secunderabad Zone - South-East Wing"),
    "appa junction": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "osmangunj": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "osmanguni": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "risala abdulla": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "kishan gunj": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "sikh village": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "sikh road": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "sikhwal": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "sitaram nagar": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "sp road": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "surya towers": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "rtc cross road": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "ashok nagar": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "gun rock enclave": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "alexander road": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "rp road": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "ghasmandi": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "ishaq colony": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "wellington": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "gandhi bomma": ("R3", "Kavadiguda", "Secunderabad Zone - South-East Wing"),
    "dilsukhnagar": ("R3", "Saroornagar", "L.B. Nagar Zone (South-East)"),
    "dilshuknagar": ("R3", "Saroornagar", "L.B. Nagar Zone (South-East)"),
    "auto nagar": ("R3", "Nacharam", "Kapra/Uppal Zone - South-East Belt"),
    "bagh hayatnagar": ("R3", "Hayathnagar", "L.B. Nagar Zone (South-East)"),
    "baghat hayatnagar": ("R3", "Hayathnagar", "L.B. Nagar Zone (South-East)"),
    # R4
    "madapur": ("R4", "Madhapur", "Serilingampally Zone - South-West Belt"),
    "madhura nagar": ("R4", "Moosapet", "Kukatpally Zone - South-West Wing"),
    "avantinagar": ("R4", "Moosapet", "Kukatpally Zone - South-West Wing"),
    "motinagar": ("R4", "Moosapet", "Kukatpally Zone - South-West Wing"),
    "idpl colony": ("R4", "Moosapet", "Kukatpally Zone - South-West Wing"),
    "sanath nagar": ("R4", "Ameerpet", "Secunderabad Zone - West Wing"),
    "sanathnagar": ("R4", "Ameerpet", "Secunderabad Zone - West Wing"),
    "panjagutta": ("R4", "Khairatabad", "Khairatabad Zone - Central-South Belt"),
    "panzer 36": ("R4", "Jubilee Hills", "Khairatabad Zone - West Wing"),
    "road no 44": ("R4", "Jubilee Hills", "Khairatabad Zone - West Wing"),
    "road no 10c": ("R4", "Jubilee Hills", "Khairatabad Zone - West Wing"),
    "road number 10c": ("R4", "Jubilee Hills", "Khairatabad Zone - West Wing"),
    "zehra nagar": ("R4", "Jubilee Hills", "Khairatabad Zone - West Wing"),
    "kakatiya hills": ("R4", "Jubilee Hills", "Khairatabad Zone - West Wing"),
    "banjarahills": ("R4", "Jubilee Hills", "Khairatabad Zone - West Wing"),
    "kavuri hills": ("R4", "Jubilee Hills", "Khairatabad Zone - West Wing"),
    "lakamsani": ("R4", "Jubilee Hills", "Khairatabad Zone - West Wing"),
    "amara jyoti": ("R4", "Jubilee Hills", "Khairatabad Zone - West Wing"),
    "lakidikapul": ("R4", "Masab Tank", "Khairatabad Zone - Central-South Belt"),
    "lakdikapul": ("R4", "Masab Tank", "Khairatabad Zone - Central-South Belt"),
    "bogulkunta": ("R4", "Khairatabad", "Khairatabad Zone - Central-South Belt"),
    "domalguda": ("R4", "Khairatabad", "Khairatabad Zone - Central-South Belt"),
    "necklace road": ("R4", "Khairatabad", "Khairatabad Zone - Central-South Belt"),
    "jalavihar": ("R4", "Khairatabad", "Khairatabad Zone - Central-South Belt"),
    "necklace pride": ("R4", "Khairatabad", "Khairatabad Zone - Central-South Belt"),
    "raj bhavan road": ("R4", "Khairatabad", "Khairatabad Zone - Central-South Belt"),
    "fateh maidan": ("R4", "Khairatabad", "Khairatabad Zone - Central-South Belt"),
    "khan lateef khan": ("R4", "Khairatabad", "Khairatabad Zone - Central-South Belt"),
    "gunfoundry": ("R4", "Khairatabad", "Khairatabad Zone - Central-South Belt"),
    "mitti ka sher": ("R4", "Charminar", "Charminar Zone - Old City (South-West)"),
    "charkaman": ("R4", "Charminar", "Charminar Zone - Old City (South-West)"),
    "lower dhoolpet": ("R4", "Charminar", "Charminar Zone - Old City (South-West)"),
    "dhoolpet": ("R4", "Charminar", "Charminar Zone - Old City (South-West)"),
    "puranapul": ("R4", "Charminar", "Charminar Zone - Old City (South-West)"),
    "mathru sree nagar": ("R4", "Madhapur", "Serilingampally Zone - South-West Belt"),
    "matrusri nagar": ("R4", "Madhapur", "Serilingampally Zone - South-West Belt"),
    "gopanpalle": ("R4", "Serilingampally", "Serilingampally Zone - South-West Belt"),
    "gopana palli": ("R4", "Serilingampally", "Serilingampally Zone - South-West Belt"),
    "nanakaramguda": ("R4", "Serilingampally", "Serilingampally Zone - South-West Belt"),
    "gowlidoddy": ("R4", "Serilingampally", "Serilingampally Zone - South-West Belt"),
    "gowlidoddi": ("R4", "Serilingampally", "Serilingampally Zone - South-West Belt"),
    "krishe sapphire": ("R4", "Serilingampally", "Serilingampally Zone - South-West Belt"),
    "ramky towers": ("R4", "Serilingampally", "Serilingampally Zone - South-West Belt"),
    "ramky": ("R4", "Serilingampally", "Serilingampally Zone - South-West Belt"),
    "knowledge capital": ("R4", "Madhapur", "Serilingampally Zone - South-West Belt"),
    "sonali spazio": ("R4", "Serilingampally", "Serilingampally Zone - South-West Belt"),
    "fabcity": ("R4", "Serilingampally", "Serilingampally Zone - South-West Belt"),
    "hardware park": ("R4", "Serilingampally", "Serilingampally Zone - South-West Belt"),
    "raheja": ("R4", "Madhapur", "Serilingampally Zone - South-West Belt"),
    "hitec city": ("R4", "Madhapur", "Serilingampally Zone - South-West Belt"),
    "mindspace": ("R4", "Madhapur", "Serilingampally Zone - South-West Belt"),
    "white fields": ("R4", "Madhapur", "Serilingampally Zone - South-West Belt"),
    "sundew": ("R4", "Madhapur", "Serilingampally Zone - South-West Belt"),
    "h06b": ("R4", "Madhapur", "Serilingampally Zone - South-West Belt"),
    "walker road": ("R4", "Masab Tank", "Khairatabad Zone - Central-South Belt"),
    "walkers road": ("R4", "Masab Tank", "Khairatabad Zone - Central-South Belt"),
    "rain center": ("R4", "Jubilee Hills", "Khairatabad Zone - West Wing"),
    "virtusa": ("R4", "Serilingampally", "Serilingampally Zone - South-West Belt"),
    "olympus": ("R4", "Serilingampally", "Serilingampally Zone - South-West Belt"),
    "sri sai towers": ("R4", "Serilingampally", "Serilingampally Zone - South-West Belt"),
    # R5 outside-GHMC (Hyderabad fringe)
    "ramoji film city": ("R5", "Hayathnagar Fringe (Abdullapurmet)", "Outside GHMC - Hayathnagar Fringe"),
    "ramoji": ("R5", "Hayathnagar Fringe (Abdullapurmet)", "Outside GHMC - Hayathnagar Fringe"),
    "sitara hotel": ("R5", "Hayathnagar Fringe (Abdullapurmet)", "Outside GHMC - Hayathnagar Fringe"),
    "anajpur": ("R5", "Hayathnagar Fringe (Abdullapurmet)", "Outside GHMC - Hayathnagar Fringe"),
    "eenadu": ("R5", "Hayathnagar Fringe (Abdullapurmet)", "Outside GHMC - Hayathnagar Fringe"),
    "tukkuguda": ("R5", "Shamshabad (Rangareddy)", "Outside GHMC (non-municipal)"),
    "kagazghat": ("R5", "Ghatkesar Belt", "Outside GHMC (non-municipal)"),
    "agapally": ("R5", "Ghatkesar Belt", "Outside GHMC (non-municipal)"),
    "zaheerabad": ("R5", "Sangareddy District", "Outside GHMC (non-municipal)"),
    "pashamylaram": ("R5", "Sangareddy District", "Outside GHMC (non-municipal)"),
    "isnapuar": ("R5", "Sangareddy District", "Outside GHMC (non-municipal)"),
    "mankhal": ("R5", "Adibatla (2026-merged)", "Outside GHMC (non-municipal)"),
    "maheshwaram": ("R5", "Adibatla (2026-merged)", "Outside GHMC (non-municipal)"),
}

ALIAS_TIGHT = {tight(k) for k in ALIASES}
for k, (route, circle, zone) in ALIASES.items():
    add(k, route, circle, zone)

known_sorted = sorted(known.items(), key=lambda kv: len(kv[0]), reverse=True)

# Hyderabad house-number prefixes that pin an address to the central-west (R4)
# belt when no named locality is present (8-2 = Banjara Hills/Jubilee Hills,
# 8-1 / 6-1 / 6-3 / 3-6 = Masab Tank / Khairatabad / Punjagutta-Somajiguda).
WEAK_PREFIX = [
    (re.compile(r"6-1-\d"), "R4", "Khairatabad", "Khairatabad Zone - Central-South Belt", "6-1-"),
    (re.compile(r"6-3-\d"), "R4", "Khairatabad", "Khairatabad Zone - Central-South Belt", "6-3-"),
    (re.compile(r"8-1-\d"), "R4", "Khairatabad", "Khairatabad Zone - Central-South Belt", "8-1-"),
    (re.compile(r"8-2-\d"), "R4", "Jubilee Hills", "Khairatabad Zone - West Wing", "8-2-"),
    (re.compile(r"3-6-\d"), "R4", "Khairatabad", "Khairatabad Zone - Central-South Belt", "3-6-"),
]

# Cities / districts outside Hyderabad -> R7. Matched on the tight form so
# punctuation never breaks it.
NON_HYD = [
    "chennai", "bengaluru", "bangalore", "bhattarahalli", "yelahanka", "varthur",
    "byatarayanapura", "belagavi", "ballari", "bellary", "mumbai", "pune", "solapur",
    "nanded", "nagpur", "kolkata", "kochi", "ernakulam", "muvattupuzha", "madurai",
    "coimbatore", "vijayawada", "visakhapatnam", "vizag", "guntur", "ongole",
    "nellore", "vedayapalem", "rajamahendravaram", "rajahmundry", "eluru",
    "china amiram", "anantapur", "kurnool", "kadapa", "tirupati", "nalgonda",
    "warangal", "karimnagar", "khammam", "mahabubnagar", "jadcherla", "medak",
    "siddipet", "mulugu", "banda thimmapur", "deoghar", "begusarai", "nagdah",
    "guwahati", "kamrup", "bahadurgarh", "jhajjar", "ludhiana", "jalandhar",
    "lucknow", "gomti nagar", "meerut", "pant nagar", "bhopal", "indore", "vadodara",
    "chhani", "bikaner", "jaipur", "raipur", "bhilai", "balasore",
    "karanjia", "dugri", "padamavati", "manish plaza", "sivaji nagar", "phoenix mall",
    "udham singh nagar", "ramateeth", "kulkarni layout", "kanbargi", "avanavathi",
    "ananthapuram", "coonoor", "nilgiris", "ooty", "kerala", "karnataka",
    "maharashtra", "tamil nadu", "tamilnadu", "uttar pradesh", "madhya pradesh",
    "gujarat", "rajasthan", "west bengal", "odisha", "assam", "bihar", "jharkhand",
    "haryana", "punjab", "chhattisgarh", "andhra pradesh",
    "rayanapadu", "woxsen", "woksen", "channasandra",
]
NON_HYD_TIGHT = {tight(x) for x in NON_HYD}

# Strong = full city/district name (enough to override a foreign-city hit).
# Bare "hyd" (e.g. a street called "HYD ROAD") is only a weak signal.
STRONG_HYD = re.compile(
    r"\bhyderabad\b|\bhyderbad\b|\bsecunderabad\b|\bsec ?bad\b|\branga ?reddy\b|\bbowenpally\b",
    re.I,
)
WEAK_HYD = re.compile(r"\bhyd\b", re.I)


def classify(row):
    name = row[3]
    addr = row[10]
    city = (row[11] or "").strip().lower()
    addr_s = "" if addr is None else str(addr)
    if not addr_s.strip():
        return "NO-ADDRESS", None, None, None, "no address in file"
    t = tight(addr_s)

    hit = None
    matched_k = None
    for k, info in known_sorted:
        if k in t:
            hit = info
            matched_k = k
            break

    non_hyd = any(x in t or (city and x in tight(city)) for x in NON_HYD_TIGHT)
    norm_addr = norm(addr_s)
    strong_hyd = bool(STRONG_HYD.search(norm_addr))
    weak_hyd = bool(WEAK_HYD.search(norm_addr))

    if non_hyd and not strong_hyd:
        m = next((x for x in NON_HYD
                  if tight(x) in t or (city and tight(x) in tight(city))), None)
        place = m.title() if m else "Outside Hyderabad"
        return "R7", None, "Other (non-Hyderabad)", place, "outside Hyderabad (%s)" % place

    if hit:
        route, circle, zone = hit
        note = "alias match" if matched_k in ALIAS_TIGHT else "matched"
        return route, circle, zone, circle, note

    if strong_hyd or weak_hyd or t:
        for rx, rte, circ, zne, label in WEAK_PREFIX:
            if rx.search(addr_s):
                return rte, circ, zne, circ, "house-number prefix %s" % label
        if strong_hyd or weak_hyd:
            return "UNMATCHED-HYD", None, "Hyderabad - review", None, "Hyderabad, area unclear"
    return "UNKNOWN", None, "Review needed", None, "incomplete address - review"


# --------------------------------------------------------------------------- #
# Load customers + classify                                                     #
# --------------------------------------------------------------------------- #
wb_src = openpyxl.load_workbook(CUSTOMERS, data_only=True)
ws_src = wb_src["Sheet1"]
cust_rows = list(ws_src.iter_rows(min_row=2, values_only=True))

ledger = []
for i, r in enumerate(cust_rows, start=2):
    route, circle, zone, area, note = classify(r)
    ledger.append({
        "sno": i - 1,
        "name": r[3],
        "addr": ("" if r[10] is None else str(r[10])).strip(),
        "route": route,
        "circle": circle,
        "area": area,
        "zone": zone,
        "note": note,
    })

ORDER_IDX = {"R1": 0, "R2": 1, "R3": 2, "R4": 3, "R5": 4, "R7": 5,
             "UNMATCHED-HYD": 6, "UNKNOWN": 7, "NO-ADDRESS": 8}
ledger_sorted = sorted(ledger, key=lambda d: (ORDER_IDX.get(d["route"], 9), d["sno"]))
by_route = Counter(d["route"] for d in ledger)

# --------------------------------------------------------------------------- #
# Write the workbook                                                            #
# --------------------------------------------------------------------------- #
wb = openpyxl.load_workbook(XLSX)
ws = wb["Routes"]
for rng in list(ws.merged_cells.ranges):
    try:
        ws.unmerge_cells(str(rng))
    except KeyError:
        pass
ws.delete_rows(2, ws.max_row)

headers = ["S. No.", "R1 = North-East (NE)", "R2 = North-West (NW)",
           "R3 = South-East (SE)", "R4 = South-West (SW)",
           "R5 = Outside GHMC (Hyderabad fringe)", "R7 = Other Routes (Outside Hyderabad)"]
for c, text in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=c, value=text)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = BOX

ws.merge_cells("A2:G2")
note = (
    "Central Hub: Old Bowenpally (Secunderabad) \u00b7 Adjacent-Route Overlap Strategy \u2014 "
    "R1 North-East \u00b7 R2 North-West \u00b7 R3 South-East \u00b7 R4 South-West \u2014 "
    "vehicles swing only into adjacent quadrants; opposite quadrants stay blocked. "
    "R5 = outside-GHMC Hyderabad fringe. R7 = other routes outside Hyderabad. "
    "All GHMC circles below (complete), then the full customer address ledger "
    "covering every row of the ALL_Customers file with no skips.")
nc = ws.cell(row=2, column=1, value=note)
nc.font = NOTE_FONT
nc.fill = NOTE_FILL
nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
for col in range(1, 8):
    ws.cell(row=2, column=col).fill = NOTE_FILL

row = 3
sno = 0
for route in ROUTE_ORDER:
    col = COL_BY_ROUTE[route]
    for zone in DATA.get(route, {"zones": []})["zones"]:
        zc = ws.cell(row=row, column=col, value=zone["zone"])
        zc.font = ZONE_FONT
        zc.fill = ZONE_FILL
        zc.alignment = BODY_ALIGN
        zc.border = BOX
        row += 1
        for circle_name, _note, _wards in zone["circles"]:
            sno += 1
            a = ws.cell(row=row, column=1, value=sno)
            a.font = BODY_FONT
            a.alignment = BODY_ALIGN
            a.border = BOX
            cc = ws.cell(row=row, column=col, value=circle_name)
            cc.font = BODY_FONT
            cc.alignment = BODY_ALIGN
            cc.border = BOX
            if sno % 2 == 0:
                a.fill = ZEBRA_FILL
                cc.fill = ZEBRA_FILL
            row += 1

# R7 column: one row per distinct non-Hyderabad city
r7_cities = sorted({(d["note"].split("(")[-1].rstrip(")").title())
                    for d in ledger if d["route"] == "R7"})
zc = ws.cell(row=row, column=COL_BY_ROUTE["R7"],
             value="R7 = Other Routes \u2014 non-Hyderabad cities served")
zc.font = ZONE_FONT
zc.fill = ZONE_FILL
zc.alignment = BODY_ALIGN
zc.border = BOX
row += 1
for city in r7_cities:
    cc = ws.cell(row=row, column=COL_BY_ROUTE["R7"], value=city)
    cc.font = BODY_FONT
    cc.alignment = BODY_ALIGN
    cc.border = BOX
    if row % 2 == 0:
        cc.fill = ZEBRA_FILL
    row += 1

top_last = row - 1

# borders on empty route cells of the populated top rows
for r in range(3, top_last + 1):
    for c in range(1, 8):
        cell = ws.cell(row=r, column=c)
        if cell.border.left is None or not cell.border.left.style:
            cell.border = BOX

# ---------------- customer address ledger ----------------
section = top_last + 2
ws.merge_cells(start_row=section, start_column=1, end_row=section, end_column=7)
sc = ws.cell(row=section, column=1,
             value="CUSTOMER ADDRESS LEDGER \u2014 ALL_Customers 31st july.xlsx "
                   "(all %d rows, complete coverage)" % len(ledger))
sc.font = BOLD_FONT
sc.fill = HEADER_FILL
sc.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
sc.alignment = HEADER_ALIGN
for c in range(1, 8):
    ws.cell(row=section, column=c).fill = HEADER_FILL

hdr_row = section + 1
ledger_headers = ["S.No.", "Customer", "Address", "Route", "Circle / Area",
                  "Planning Zone", "Match Note"]
for c, text in enumerate(ledger_headers, start=1):
    cell = ws.cell(row=hdr_row, column=c, value=text)
    cell.font = BOLD_FONT
    cell.fill = LEDGER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = BOX

rr = hdr_row + 1
for i, d in enumerate(ledger_sorted):
    vals = [d["sno"], d["name"], d["addr"], d["route"],
            d["area"] or "\u2014", d["zone"] or "\u2014", d["note"]]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=rr, column=c, value=v)
        cell.font = BODY_FONT
        cell.border = BOX
        if c == 1:
            cell.alignment = CENTER_ALIGN
        else:
            cell.alignment = BODY_ALIGN
        if i % 2 == 1:
            cell.fill = ZEBRA_FILL
    if d["route"] in ("UNKNOWN", "UNMATCHED-HYD", "NO-ADDRESS"):
        for c in range(1, 8):
            ws.cell(row=rr, column=c).fill = REVIEW_FILL
    rr += 1
last = rr - 1

ws.freeze_panes = "A3"
ws.auto_filter.ref = "A1:G%d" % top_last
ws.sheet_view.showGridLines = True

for col, width in (("A", 7), ("B", 30), ("C", 34), ("D", 34),
                   ("E", 34), ("F", 34), ("G", 30)):
    ws.column_dimensions[col].width = width

wb.save(XLSX)

# --------------------------------------------------------------------------- #
# Customer coverage JSON (mirrors the workbook ledger; includes R7)            #
# --------------------------------------------------------------------------- #
coverage = {
    "meta": {
        "title": "SkyLimit Customer Route Coverage \u2014 ALL_Customers 31st july.xlsx",
        "hub": "Old Bowenpally",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "Customers/ALL_Customers 31st july.xlsx (Sheet1)",
        "customer_count": len(ledger),
    },
    "summary": {
        "R1": by_route.get("R1", 0),
        "R2": by_route.get("R2", 0),
        "R3": by_route.get("R3", 0),
        "R4": by_route.get("R4", 0),
        "R5": by_route.get("R5", 0),
        "R7": by_route.get("R7", 0),
        "UNMATCHED-HYD": by_route.get("UNMATCHED-HYD", 0),
        "UNKNOWN": by_route.get("UNKNOWN", 0),
        "NO-ADDRESS": by_route.get("NO-ADDRESS", 0),
    },
    "r7": {
        "count": by_route.get("R7", 0),
        "cities": r7_cities,
    },
    "records": [
        {
            "sno": d["sno"],
            "customer": d["name"],
            "address": d["addr"],
            "route": d["route"],
            "area": d["area"],
            "zone": d["zone"],
            "note": d["note"],
        }
        for d in ledger_sorted
    ],
}
coverage_path = REPO / "ghmc-routing" / "output" / "customers_route_map.json"
coverage_path.write_text(
    json.dumps(coverage, indent=2, ensure_ascii=False), encoding="utf-8"
)

print("Canonical circles: %d" % sno)
print("Ledger rows:       %d" % len(ledger))
print("By route:")
for route in ["R1", "R2", "R3", "R4", "R5", "R7",
              "UNMATCHED-HYD", "UNKNOWN", "NO-ADDRESS"]:
    print("  %-13s %d" % (route, by_route.get(route, 0)))
print("Rows: header(1) + note(1) + top(%d..%d) + ledger(%d..%d)"
      % (3, top_last, hdr_row + 1, last))
print("Coverage JSON: %s" % coverage_path.name)
